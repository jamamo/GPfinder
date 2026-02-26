"""
One-time script: imports GP data from Excel files into gp.db
Run once after creating the database: py -3 import_gps.py
"""
import os
import sqlite3
import openpyxl
from werkzeug.security import generate_password_hash

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, 'gp.db')
EXCEL_DIR  = r'C:\Users\Owner\Desktop\GPlists'


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db(conn):
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS gps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            practice_code TEXT,
            practice_name TEXT NOT NULL,
            partnership_name TEXT,
            neighbourhood TEXT,
            area TEXT,
            address_line1 TEXT,
            address_line2 TEXT,
            address_line3 TEXT,
            postcode TEXT,
            telephone TEXT,
            email TEXT,
            region TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        );
    ''')
    if not conn.execute('SELECT id FROM admins LIMIT 1').fetchone():
        conn.execute(
            'INSERT INTO admins (username, password_hash) VALUES (?, ?)',
            ('admin', generate_password_hash('admin123'))
        )
    conn.commit()


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def insert(conn, row):
    conn.execute('''
        INSERT INTO gps
          (practice_code, practice_name, partnership_name, neighbourhood, area,
           address_line1, address_line2, address_line3, postcode, telephone, email, region)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    ''', row)


def import_manchester(conn, path):
    """
    Manchester GPs.xlsx
    Columns (0-indexed, no header row):
      0  = (empty)
      1  = (empty / locality grouping)
      2  = Neighbourhood
      3  = Practice Code
      4  = P Code (duplicate)
      5  = Practice Name
      6  = Address line 1
      7  = Full address string (skip)
      8  = Postcode
      9  = Telephone Number
      10 = Generic Email
      11 = Practice Manager (skip)
      12 = Practice Manager Email (skip)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(values_only=True):
        name = clean(row[5]) if len(row) > 5 else None
        if not name:
            continue
        code = clean(row[3]) if len(row) > 3 else None
        insert(conn, (
            code,                                   # practice_code
            name,                                   # practice_name
            None,                                   # partnership_name
            clean(row[2]) if len(row) > 2 else None,  # neighbourhood
            'Manchester',                           # area
            clean(row[6]) if len(row) > 6 else None,  # address_line1
            None,                                   # address_line2
            None,                                   # address_line3
            clean(row[8]) if len(row) > 8 else None,  # postcode
            clean(row[9]) if len(row) > 9 else None,  # telephone
            clean(row[10]) if len(row) > 10 else None, # email
            'Manchester',                           # region
        ))
        count += 1
    print(f'  Manchester: {count} records')
    return count


def import_preston(conn, path):
    """
    Preston & Chorley GPs.xlsx
    2 blank rows, then header at row 3, data from row 4.
    Columns: CCG | P/Code | Practice Name | Address 1 | Address 2 | Address 3 | Post Code | Tel No | Generic Email
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_found = False
    count = 0
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            # Detect header row by looking for 'Practice Name'
            row_vals = [str(v).strip() if v else '' for v in row]
            if 'Practice Name' in row_vals:
                header_found = True
            continue
        name = clean(row[2]) if len(row) > 2 else None
        if not name:
            continue
        insert(conn, (
            clean(row[1]) if len(row) > 1 else None,   # practice_code
            name,                                        # practice_name
            None,                                        # partnership_name
            None,                                        # neighbourhood
            clean(row[0]) if len(row) > 0 else None,   # area (CCG)
            clean(row[3]) if len(row) > 3 else None,   # address_line1
            clean(row[4]) if len(row) > 4 else None,   # address_line2
            clean(row[5]) if len(row) > 5 else None,   # address_line3
            clean(row[6]) if len(row) > 6 else None,   # postcode
            clean(row[7]) if len(row) > 7 else None,   # telephone
            clean(row[8]) if len(row) > 8 else None,   # email
            'Preston & Chorley',                        # region
        ))
        count += 1
    print(f'  Preston & Chorley: {count} records')
    return count


def import_stockport(conn, path):
    """
    Stockport GP Practice_generic email addresses_4.10.2019.xlsx
    Header at row 1:  Network | P Code | Partnership Name | Practice Name | Practice Email Address
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_skipped = False
    count = 0
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        name = clean(row[3]) if len(row) > 3 else None
        if not name:
            continue
        insert(conn, (
            clean(row[1]) if len(row) > 1 else None,   # practice_code
            name,                                        # practice_name
            clean(row[2]) if len(row) > 2 else None,   # partnership_name
            None,                                        # neighbourhood
            clean(row[0]) if len(row) > 0 else None,   # area (Network)
            None,                                        # address_line1
            None,                                        # address_line2
            None,                                        # address_line3
            None,                                        # postcode
            None,                                        # telephone
            clean(row[4]) if len(row) > 4 else None,   # email
            'Stockport',                                 # region
        ))
        count += 1
    print(f'  Stockport: {count} records')
    return count


def main():
    conn = get_db()
    init_db(conn)

    existing = conn.execute('SELECT COUNT(*) FROM gps').fetchone()[0]
    if existing > 0:
        ans = input(f'Database already has {existing} records. Re-import? [y/N] ').strip().lower()
        if ans != 'y':
            print('Aborted.')
            conn.close()
            return
        conn.execute('DELETE FROM gps')
        conn.commit()

    files = {f.lower(): f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')}

    total = 0
    print('Importing...')

    # Manchester — prefer the original (no "Copy of" prefix)
    mch = next((f for k, f in files.items()
                 if 'manchester' in k and not k.startswith('copy')), None)
    if mch:
        total += import_manchester(conn, os.path.join(EXCEL_DIR, mch))
    else:
        print('  Manchester: file not found')

    # Preston & Chorley — prefer original
    pre = next((f for k, f in files.items()
                 if 'preston' in k and not k.startswith('copy')), None)
    if pre:
        total += import_preston(conn, os.path.join(EXCEL_DIR, pre))
    else:
        print('  Preston & Chorley: file not found')

    # Stockport — prefer original
    sto = next((f for k, f in files.items()
                 if 'stockport' in k and not k.startswith('copy')), None)
    if sto:
        total += import_stockport(conn, os.path.join(EXCEL_DIR, sto))
    else:
        print('  Stockport: file not found')

    conn.commit()
    conn.close()
    print(f'\nDone. {total} GP practices imported.')
    print('Run the app with:  py -3 app.py')


if __name__ == '__main__':
    main()
